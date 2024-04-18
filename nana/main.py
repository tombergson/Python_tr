import openpyxl

inv_file = openpyxl.load_workbook("nana/inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}
total_value_per_supp = {}
products_under_10_inv = {}


for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

   # calulation for number of products per supplier
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier[supplier_name] 
        products_per_supplier[supplier_name] = products_per_supplier[supplier_name] + 1
    
    else:
        
        products_per_supplier[supplier_name] = 1 


    # calculation total value per supplier
    if supplier_name in total_value_per_supp:
        current_total_value = total_value_per_supp.get(supplier_name)
        total_value_per_supp[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supp[supplier_name] = inventory * price

   # logic product with inventory less than 10
    if inventory < 10:
        products_under_10_inv[int(product_num)] = int(inventory)
     

    #add total inventory price
    inventory_price.value = inventory * price


print(products_per_supplier) 
print(total_value_per_supp) 
print(products_under_10_inv) 


inv_file.save("nana/inventory_with_total_value.xlsx")
